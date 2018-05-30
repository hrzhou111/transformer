# -*- coding: utf-8 -*-
"""
Created on Sun Mar 25 25:24:43 2018

@author: zhr
"""

import gdspy
import numpy
import openpyxl as opx

class Transformer:
    """
    图层说明：
    layer function material thickness
    1 L1 Nb 100nm
    2 Sio2 绝缘
    3 电阻层 Mo 2Ω/方块
    4 Sio2 绝缘
    5 washer Nb 100nm
    
    Nb 薄膜
    L1:SNSPD端  电感  多线圈
    条件：L1>L2
    线圈总长度：l = 4Ni[L1d+(Ni+1)si+Ni*wi]   方形线圈
    Lstripe=u0*(h+lambdaLw+lambda_Li)/(wi+2(h+lambdaLw+lambda_Li))
    h 绝缘层的薄膜厚度，wi 线圈的线宽，   lambda_Lw  lambda_Li 输入线圈和washer的伦敦穿透深度
    L1 = Ni^2(Lh+Lsl/3)+lstripe*l
    L2 ：SFQ端电感 单圈线圈
     L = Lh+Lsl+LJp
     Lh = 1.25u0*L2d
     Lsl = 0.3Ph/μm
     Ljp = 0
    M：互感
    M = Ni*(Lh+Ll/2)  L1端的参数
    单位：nH，μm  
    常数值：u0 = 4*pi*10^-7Vs/Am
      lambda_Li = lambda_lw = 90nm
     precision：工艺精度，单位μm
    """
    Ni = 0
    d = 0
    h = 0.1
    L2w = 20
    L1 = 0
    L2 = 0
    M = 0
    k = 0
    L1wi = 0
    L1si = 0
    lambda_Lw = 0.085    #Nb薄膜伦敦穿透深度85nm
    lambda_Li=0.085
    plottype =0
    precision = 1  #工艺加工精度，单位um
    def  __init__(self,text,L1,L2,L1wi,L1si,L1h,lambda_Li,amp,rsquire,r=50,precision=1):

        """
        L1,L2 等效电感，二端口网络的等效电感
        """
        self.text = str(text)
        self.L1wi = L1wi
        self.L1si = L1si
        self.precision = precision
        self.u0 = 4*numpy.pi*10**-7*10**9/10**6  #单位 nH/μm 
        self.rsquire = rsquire
        self.r = r
        Lmin =self.u0*(2*precision)*1.25+0.0003*2*precision   #L2washer应该大于这个值
        if L2<Lmin/2:
            print("L2电感过小,不能画出版图")
            return
        self.L2 = L2*2
        self.L1 = L1/2
        self.plottype = 2  #dual washer
        self.d = self.L2/(1.25*self.u0+0.0006)   #强制狭缝宽度=2d
        self.L2w = max(2*self.d,0.1*L2/0.0003)   
        Lsl = 0.0003*self.L2w
        self.d = (self.L2-Lsl)/1.25/self.u0  #方形线圈的直径
        print("L1{}nH,L2{}nH".format(L1,L2))
         #求解Ni一元二次方程的值
        self.lstripe = self.u0*(L1h+self.lambda_Lw+lambda_Li)/(L1wi+2*(L1h+self.lambda_Lw+lambda_Li))  #微带线单位长度电感 nH/um
        Lh = 1.25*self.u0*self.d   # SFQ端为单层线圈
        a = Lh+Lsl/3+4*(L1wi+L1si)*self.lstripe
        b = 4*(L1si+self.d)*self.lstripe
        c = - self.L1
        delta = numpy.sqrt(b**2-4*a*c)
        Ni1 = (-1*b-delta)/2/a
        Ni2 = (-1*b+delta)/2/a
        if Ni2<0:
             print("参数设置不合理，无法画出版图")
             return
         
        Ni = int(Ni2)
        if self.Ni-Ni2>0.8:
            Ni = int(Ni)+1
        self.Ni = Ni
        self.L2w = min(2*self.d,self.Ni*(self.L1wi+self.L1si))+5
        self.d = (self.L2-0.0003*self.L2w)/self.u0/1.25
        self.slite = 25
        if self.caluncertain()['L1']-self.L1*self.plottype<-0.1*self.L1*self.plottype:
            self.Ni = self.Ni+1
            self.L2w = max(2*self.d,self.Ni*(self.L1wi+self.L1si))-1
            self.d = (self.L2-0.0003*self.L2w)/self.u0/1.25
            

        self.M = self.Ni*(Lh+Lsl/2)
        self.k = self.M/numpy.sqrt(self.L1*self.L2)
        

#        self.elec_width = elec_width

    def plottransformer(self):
        print(self.d,self.Ni,self.L1,self.L2,self.M)
        cellname = 'L1_{}nH_L2_{}nH,M{}nH'.format(self.L1,self.L2,self.M)
        cell = gdspy.Cell(cellname,exclude_from_current=True)

        #绘制sspd端输入线圈 Ni圈，电感值L1，位于底层，layer1
        if self.plottype==2:
            #电阻层 layer = 3 电阻
            layer = {'layer':3,'datatype':3}
#            Nsquire = self.r/self.rsquire + 2  #2表示接触窗口，不算电阻，
#            startposition = (slite/2+self.L2w+self.d/2-self.d/2-self.L1wi-self.Ni*(self.L1wi+self.L1si),
#                             -self.d/2-self.L1wi/2-self.Ni*(self.L1wi+self.L1si))
#            start = (-slite/2,0)
#            width = self.L1wi+2*self.precision+self.L1wi*3
            unitsquire = self.L1wi

            if(self.r<self.rsquire):
                lengthr = unitsquire+6*unitsquire
                widthr = self.rsquire*unitsquire/self.r
            else:
                widthr = unitsquire
                lengthr = unitsquire*self.r/self.rsquire+6*unitsquire
                
            startr = (0,lengthr/2)
                
            R = gdspy.Path(widthr,startr).segment(lengthr,'-y',**layer)  #电阻线
            cell.add(R)
            
            
              #washer 层
            self.slite = (self.L1wi+self.L1si)*2+widthr+2*self.precision
            slite = self.slite
            d = self.d
            self.d = self.d-1
            
            layer = {'layer':5,'datatype':5}
            startposition = (self.d/2+self.L2w/2-(self.L2w+self.d/2+slite/2+0.5),self.L1wi/2+self.precision+5)
            turn = [1]*4
            turn.extend([-10/self.L2w,-1])
            turn.extend([1]*4)
            turn.append(-10/self.L2w)
            
            length = [self.d/2+self.L2w/2-(self.L1wi/2+self.precision+5),self.d+self.L2w,
                  self.d+self.L2w,self.d+self.L2w,self.d/2+self.L2w/2-(self.L1wi/2+self.precision+5),
                  self.L2w+slite+1,
                  self.d/2+self.L2w/2-(self.L1wi/2+self.precision+5),self.d+self.L2w,
                  self.d+self.L2w,self.d+self.L2w,self.d/2+self.L2w/2-(self.L1wi/2+self.precision+5),
                  self.L2w*1.5+slite+1]
            L2path = gdspy.L1Path(startposition,'+y',self.L2w,length,turn,**layer)
            L2path = gdspy.fast_boolean(L2path,None,'or',**layer)    #washer
            self.d = d
#            L2path.translate(0.5,0.5)
#           L2path.fillet(self.L2w/4,points_per_2pi = 100,max_points = 30)
            
            #input coil
            startposition = (-(self.L2w+self.d/2+slite/2)+self.d/2+self.L1wi/2,-self.L1wi/2)
            turn = [1]*(self.Ni*4)
            turn.pop(0)
            turn.append(-1)
            turn.append(-1)
            length = []
            for i in range(self.Ni):
                length.extend([self.d+self.L1wi+(2*i-1)*(self.L1wi+self.L1si),
                           self.d+self.L1wi+2*i*(self.L1wi+self.L1si),
                           self.d+self.L1wi+2*i*(self.L1wi+self.L1si),
                           self.d+self.L1wi+(2*i+1)*(self.L1wi+self.L1si)])
            length[0] = self.d/2+self.L1wi
            length.append(self.d+self.L1wi*1.5+(2*self.Ni-1)*(self.L1wi+self.L1si))
            length.append(self.d+self.L1wi*1.5+(2*self.Ni-1)*(self.L1wi+self.L1si))
            layer={'layer':1,'datatype':1} 
            L1path = gdspy.L1Path(startposition,'+y',self.L1wi,length,turn,**layer)
#            L1path.fillet(self.L1si/10,points_per_2pi=30, max_points=30)
#            cell.add(L1path)
            L11 = L1path.polygons
            L1_2 = gdspy.PolygonSet(L11,**layer).rotate(numpy.pi,(0,0))
#            cell.add(L1_2)
    
            
            
            
            
            
            #washer 层和电阻层并联

#            cell.add(link_R_w2)
#            L2path.translate(0.5,0.5)
            
         #电阻和wsher连接层
            layer = {'layer':4,'datatype':4}
            mid_r_w = gdspy.fast_boolean(R,L2path,'and',**layer)
            temp = []
            layer = {'layer':4,'datatype':4}
            for j in mid_r_w.polygons:
                temp.append(self.amplifier(self.get_center(j),j,0.8))
            cell.add(gdspy.PolygonSet(temp,**layer))
            temp = []
            layer = {'layer':2,'datatype':2}
            startNb = (startr[0]-widthr/2-0.5,unitsquire*1.5)
            Nb_link = gdspy.Path(2*unitsquire,startNb)
            Nb_link.segment(widthr+1,'+x',**layer)
#            Nb_link = gdspy.Rectangle((-widthr/2-0.5,0.5*unitsquire),(widthr/2+0.5,2.5*unitsquire),**layer)
            cell.add(Nb_link)
            Nb_link2 = self.symetry('x',*Nb_link.polygons)
            cell.add(gdspy.PolygonSet([Nb_link2],**layer))
            layer = {'layer':1,'datatype':1}
            Nb_cover = gdspy.PolygonSet([self.amplifier(self.get_center(*Nb_link.polygons),*Nb_link.polygons,1+0.5/self.L1wi)],**layer)
            Nb_cover2 = self.symetry('x',*Nb_cover.polygons)
            Nb_cover2 = gdspy.PolygonSet([Nb_cover2],**layer)
            cell.add(Nb_cover)
            cell.add(Nb_cover2)
#            
                

            

            #input线圈washer层连接线
            wi = self.L1wi
            self.L1wi = self.L1wi+2
            layer = {'layer':5,'datatype':5}
            washer_mid = gdspy.Rectangle((-(self.L2w+self.d/2+slite/2+1)+self.d/2,-self.L1wi/2),
                                     ((self.L2w+self.d/2+slite/2+1)-self.d/2,self.L1wi/2),**layer)  
            washer_mid = gdspy.Polygon(washer_mid.points,**layer)
            cell.add(washer_mid)
            self.L1wi = wi
            L2_in = gdspy.Path(4*self.L1wi,
                               (0,self.L1wi/2+self.precision+5)).segment(150-(self.L1wi/2+self.precision+5)+1.5*self.L1wi,'+y',**layer)
            L2_out = L2_in.polygons
            L2_out = gdspy.PolygonSet(L2_out,**layer).rotate(numpy.pi)
            L2path = gdspy.fast_boolean(L2path,L2_in,'or',**layer)    #washer
            L2path = gdspy.fast_boolean(L2path,L2_out,'or',**layer)    #washer
            cell.add(L2path)
            self.L1wi = wi

            
            
            #引出电极  总大小       5mmX1.5mm
            layer = {'layer':1,'datatype':1}
            boundary = 2*(self.L1si+self.L1wi)*self.Ni+self.d+slite/2+100
            cell.add(gdspy.Rectangle((boundary,50),(boundary+1000,550),**layer))
            cell.add(gdspy.Rectangle((boundary,-50),(boundary+1000,-550),**layer))
            cell.add(gdspy.Rectangle((-boundary,50),(-boundary-1000,550),**layer))
            cell.add(gdspy.Rectangle((-boundary,-50),(-boundary-1000,-550),**layer))
            L2_in_elec = gdspy.Path(3*self.L1wi,
                               (self.L1wi*2.5,150-self.L1wi/2)).segment(5*self.L1wi,'-x',**layer).segment(boundary,'-x',40*self.L1wi,**layer)
            cell.add(L2_in_elec)
            L2_out_elec = gdspy.PolygonSet(L2_in_elec.polygons,**layer).rotate(numpy.pi)
            cell.add(L2_out_elec)
            
            start = (-1*(slite/2+self.d+2*self.Ni*(self.L1wi+self.L1si))+self.L1wi+self.L1si,-self.d/2-(self.L1si+self.L1wi)*self.Ni-length[-1]+self.L1wi+self.L1si*0.5)
            lengthin = boundary-80
            L1_in_elec = gdspy.Path(self.L1wi,start).segment(lengthin,'-x',20*self.L1wi)
 
            L1path = gdspy.fast_boolean(L1path,L1_in_elec,'or',**layer)
            cell.add(L1path)
#            r_end = ((slite/2+self.L2w+self.d/2)-(self.d/2+self.Ni*(self.L1si+self.L1wi)+self.L1wi/2)+(Nsquire-1)*self.L1wi,
#                     -self.d/2-self.Ni*(self.L1si+self.L1wi))
#            length = [375-self.d/2-self.L1wi/2-self.Ni*(self.L1wi+self.L1si),
#                     boundary-(slite/2+self.L2w+self.d/2)+(self.d/2+self.Ni*(self.L1wi+self.L1si)+self.L1wi/2)-((Nsquire-1)*self.L1wi)]
#            L1_out_elec = gdspy.L1Path(r_end,'-y',self.L1wi,length,[1],**layer)
            L1_out_elec = gdspy.PolygonSet(L1_in_elec.polygons,**layer).rotate(numpy.pi)
            L1_2 = gdspy.fast_boolean(L1_2,L1_out_elec,'or',**layer)
            cell.add(L1_2)
            
            
            #引线窗口
            #L1 washer ceng 连接
            k = 0
            dis = [1,3]

            for i in[2,4]:
                layer = {'layer':i,'datatype':i}   #绝缘层
                L1link = gdspy.Rectangle((-(self.L2w+self.d/2+slite/2)+self.d/2+self.L1wi/2-((self.L1wi+2)/2)*(1-dis[k]/(self.L1wi+2)),-(self.L1wi+2)/2*(1-dis[k]/(self.L1wi+2))),
                                         (-(self.L2w+self.d/2+slite/2)+self.d/2+self.L1wi/2+1+((self.L1wi+2)/2)*(1-dis[k]/(self.L1wi+2)),(self.L1wi+2)/2*(1-dis[k]/(self.L1wi+2))),**layer)
                k = k+1
                cell.add(L1link)
                l2 = gdspy.Polygon(L1link.points,**layer)
                l2.rotate(numpy.pi)
                cell.add(l2)
            #washer层和电极接线，layer = 2 4
            for i in [2,4]:
                layer = {'layer':i,'datatype':i}   #绝缘层
                link_washer_elec = gdspy.fast_boolean(L2path,L2_out_elec,'and',**layer)
                link_washer = []
                for j in link_washer_elec.polygons:
                    center  = self.get_center(j)
                    k = self.amplifier(center,j,0.85**(5-i))
                    link_washer.append(k)
                cell.add(gdspy.PolygonSet(link_washer,**layer))
                cell.add(gdspy.PolygonSet(link_washer,**layer).rotate(numpy.pi))
            #L2与R并联
#            layer = {'layer':4,'datatype':4}
#            L2_r_link = gdspy.fast_boolean(L2path,R,'and',**layer)
#            cell.add(L2_r_link)
            #L1-2与R接口
            '''
            layer = {'layer':2,'datatype':2}
            L1_r_link = gdspy.fast_boolean(L1_2,R,'and',**layer)
            cell.add(L1_r_link)
            
            #R与L1 out 端接口
            L1_outelec_r_link = gdspy.fast_boolean(R,L1_out_elec,'and',**layer)
            cell.add(L1_outelec_r_link)
            '''
        # mark
        
        # text
        layer = {'layer':1,'datatype':1}
        cell.add(gdspy.Text(self.text,100,(-boundary-100,20),angle = -numpy.pi/2,**layer))
        
        lib = gdspy.GdsLibrary(name='lib')
        lib.add(cell,overwrite_duplicate=True)
        lib.write_gds('outfile.gds',unit = 1e-06,precision=1e-09)
        return cell
    def caluncertain(self):
        """
        计算实际电感值
        """
        L2 = (self.u0*1.25*self.d+0.0003*(self.L2w+self.slite/2))/self.plottype
        l = 4*self.Ni*(self.d+(self.Ni+1)*self.L1si+self.Ni*self.L1wi)
        L1 = self.plottype*(self.Ni**2*(self.u0*1.25*self.d+0.0003*self.L2w/3)+l*self.lstripe)
        M = self.Ni*(1.25*self.u0*self.d+0.0003*self.L2w/2)
        amp = M/L2
        return{'L1':L1,'L2':L2,'M':M,'amp':amp,'Ni':self.Ni}
    def get_center(self,points):
        x = 0
        y=0
        for point in points:
           x = x+point[0]/len(points)
           y = y+point[1]/len(points)
        return (x,y)
    def amplifier(self,center,polygons,amplifier):
        '''
        center :points like(x,y)
        '''
        polygon = []
        for j in polygons:
            polygon.append([(j[0]-center[0])*amplifier+center[0],(j[1]-center[1])*amplifier+center[1]])
    
        return numpy.array(polygon)
    def symetry(self,symline,polygons):
        '''
        单个多边形的对称变换
        支持x，y轴变换
        '''
        polygon = []
        for j in polygons:
            if symline =='x':
                polygon.append([j[0],0-j[1]])
            elif symline=='y':
                polygon.append([0-j[0],j[1]])
        return numpy.array(polygon)
            
            
        
        
    
    
def transformer_array(dx,dy,m,n,cells):
    """
    绘制纳米线阵列
    """

    arrayname = 'array{}X{}'.format(m,n)
    refcell = gdspy.Cell(arrayname,exclude_from_current=True)
    lib = gdspy.GdsLibrary(name='lib')
#    lib.add(cells,overwrite_duplicate=True)   #2017102525：25
    j = m-1
    k = n
    for i in cells:
        refcell.add(gdspy.CellReference(i,rotation = 180).translate(dx*(k-1),dy*j))
        k = k-1
        if k%(n) ==0:
            k = n
            j = j-1
    refcell.flatten()      #
    refcell1 = gdspy.Cell(arrayname,exclude_from_current=True)
    refcell1.add(gdspy.CellReference(refcell).translate(-1*dx*(n-1)/2,-1*dy*(m-1)/2))
    refcell1.flatten()
    lib.add(refcell1,overwrite_duplicate=True)
    return lib
 

def get_param_from_xlsx(filename):
    """
    从xlsx文件中读取纳米线参数设定
    param1[[],]  单独纳米线结构参数
    param2 :[]   纳米线阵列参数
    """
    wb = opx.load_workbook(filename)
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    param1 = []
    for rown in range(2,sheet.max_row+1):
        param1.append([])
        for coln in range(1,sheet.max_column+1):
            param1[rown-2].append(sheet.cell(row = rown,column = coln).value)
    param2 = []
    sheet = wb.get_sheet_by_name(sheetnames[1])
    for coln in range(1,sheet.max_column+1):
        param2.append(sheet.cell(row = 2,column = coln).value)
    wb.close()
    print(param1)

    return param1,param2

def plot(filename):
    param1,param2 = get_param_from_xlsx(filename)
    cells=[Transformer(*i).plottransformer() for i in param1]
    param2.append(cells)
    lib = transformer_array(*param2)
    dest = str(filename.replace('.xlsx','.gds'))
    lib.write_gds(dest,unit = 1e-06,precision=1e-09)
    return

if __name__ == '__main__':
    c1 = Transformer('1', 1, 0.012, 3, 3, 0.1, 0.09, 10, 2, 0.3, 3)
#  text,L1,L2,L1wi,L1si,L1h,lambda_Li,amp,rsquire,r=50,precision=1)
    c1.plottransformer()
    print(c1.caluncertain())

